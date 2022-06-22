from cgitb import reset
import openpyxl
import sqlite3

#create a workbook and a sheet named 'Violations Types'
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Violations Types'

connection = sqlite3.connect("company.db")
cursor = connection.cursor()

#query the number of each type of violation based on violation_code
sql = """ SELECT DISTINCT violation_code, violation_description, count(violation_code)
         FROM Violations
         GROUP BY violation_code
         ORDER BY violation_code;"""
cursor.execute(sql)
violations = cursor.fetchall()

#built three column names
ws['A1'].value = 'Code'
ws['B1'].value = 'Description'
ws['C1'].value = 'Count'

#loop the query result and write the data into cells
result = []
for i in range(len(violations)):
    ws.cell(row=i+2, column=1).value = violations[i][0]
    ws.cell(row=i+2, column=2).value = violations[i][1]
    ws.cell(row=i+2, column=3).value = violations[i][2]
    result.append(violations[i][2])
    total = sum(result[0:len(result)])

#write the last line in the excel file
ws.cell(row=len(violations)+2, column=2).value = 'Total Violations'
ws.cell(row=len(violations)+2, column=3).value = total

#save the workbook named 'ViolationTypes.xlsx'    
wb.save('ViolationTypes.xlsx')