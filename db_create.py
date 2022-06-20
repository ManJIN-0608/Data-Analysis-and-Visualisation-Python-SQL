import openpyxl
import sqlite3

#open two .xlsx files
wb1 = openpyxl.load_workbook('inspections.xlsx')
wb2 = openpyxl.load_workbook('violations.xlsx')
# 获取文件的表单
ws1 = wb1['inspections']
ws2 = wb2['violations']

#loop each lines in inspections.xlsx as tuples in a list
ws1_cells = []
for rows in ws1.iter_rows(min_row = 2, max_col = 20, max_row = 191372):
    row_cells = []
    for cell in rows:
        row_cells.append(str(cell.value))
    ws1_cells.append(tuple(row_cells))

#loop each lines in violations.xlsx as tuples in a list
ws2_cells = []
for rows in ws2.iter_rows(min_row = 2, max_col = 5,max_row = 906015):
   row_cells = []
   for cell in rows:
        row_cells.append(str(cell.value))
   ws2_cells.append(tuple(row_cells))

connection = sqlite3.connect('company.db')
cursor = connection.cursor()

#create table Inspections
cursor.execute("""DROP TABLE IF EXISTS Inspections""")
sql = """CREATE TABLE Inspections(
    activity_date DATE,
    employee_id VARCHAR(20),
    facility_address VARCHAR(32),
    facility_city VARCHAR(32),
    facility_id VARCHAR(20),
    facility_name VARCHAR(32),
    facility_state VARCHAR(10),
    facility_zip VARCHAR(20),
    grade CHAR(1),
    owner_id VARCHAR(20),
    owner_name VARCHAR(32),
    pe_description VARCHAR(32),
    program_element_pe CHAR(4),
    program_name TEXT,
    program_status VARCHAR(32),
    record_id VARCHAR(20),
    score INTEGER,
    serial_number VARCHAR(20),
    service_code INTEGER,
    service_description VARCHAR(32));"""
cursor.execute(sql)

#transfer data from the list in table Inspections
for p in ws1_cells:
    format_str = """INSERT INTO Inspections(
    activity_date, 
    employee_id, 
    facility_address, 
    facility_city, 
    facility_id, 
    facility_name, 
    facility_state,
    facility_zip,
    grade,
    owner_id,
    owner_name,
    pe_description,
    program_element_pe,
    program_name,
    program_status,
    record_id,
    score,
    serial_number,
    service_code,
    service_description)
    VALUES(
    "{a}", 
    "{b}", 
    "{c}", 
    "{d}", 
    "{e}",
    "{f}", 
    "{g}",
    "{h}",
    "{i}",
    "{g}",
    "{k}",
    "{l}",
    "{m}",
    "{n}",
    "{o}",
    "{p}",
    "{q}",
    "{r}",
    "{s}",
    "{t}");"""
    sql = format_str.format(
    a = p[0], 
    b = p[1], 
    # 把双引号替换成单引号
    c = p[2].replace("\"","\"\""), 
    d = p[3], 
    e = p[4],
    f = p[5].replace("\"","\"\""), 
    g = p[6],
    h = p[7],
    i = p[8],
    j = p[9],
    k = p[10].replace("\"","\"\""),
    l = p[11],
    m = p[12],
    n = p[13].replace("\"","\"\""),
    o = p[14],
    p = p[15],
    q = p[16],
    r = p[17],
    s = p[18],
    t = p[19])
    cursor.execute(sql)

connection.commit()

#create table Violations
cursor.execute("""DROP TABLE IF EXISTS Violations""")
sql = """CREATE TABLE Violations(
    points INTEGER,
    serial_number VARCHAR(20),
    violation_code VARCHAR(32),
    violation_description VARCHAR(100),
    violation_status VARCHAR(20));"""
cursor.execute(sql)

#transfer data from the list in table Violations
for q in ws2_cells:
    format_str = """INSERT INTO Violations(
    points,
    serial_number,
    violation_code,
    violation_description,
    violation_status)
    VALUES(
    "{a}", 
    "{b}", 
    "{c}", 
    "{d}", 
    "{e}");"""
    sql = format_str.format(
    a = q[0], 
    b = q[1], 
    c = q[2], 
    d = q[3], 
    e = q[4])
    cursor.execute(sql)

connection.commit()

#update and close the database
connection.close()